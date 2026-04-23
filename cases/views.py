from urllib import request
import uuid

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import PermissionDenied
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views import View
from django.views.generic import DetailView, ListView, TemplateView
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from accounts.models import User
from accounts.views import get_type_label
from .forms import BitacoraEntryForm, CaseDistributionForm
from .models import BitacoraEntry, Case, CaseAssignment
from .services import (
    CaseCompletionValidator,
    CaseDistributionService,
    CaseDocumentService,
    CaseStageManager,
    asignar_caso_automatico,
    CaseAutomaticDistributionService,
)


def get_topbar_user_data(user):
    return {
        "full_name": user.full_name,
        "role_name": user.get_role_display(),
        "initials": user.initials,
    }


class SecretaryCasesView(LoginRequiredMixin, TemplateView):
    template_name = "cases/secretary_cases.html"

    def dispatch(self, request, *args, **kwargs):
        if request.user.role != User.Role.SECRETARIA:
            messages.error(request, "No autorizado.")
            return redirect("dashboard_redirect")
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        cases = Case.objects.all().order_by("-created_at")

        context["user_data"] = get_topbar_user_data(self.request.user)
        context["cases"] = cases
        context["cases_count"] = cases.count()
        return context


class CaseReportView(LoginRequiredMixin, TemplateView):
    """
    Vista para mostrar un reporte de todos los casos del consultorio jurídico.
    Accesible para Secretaria y Asesor.
    Permite descargar los datos en Excel.
    """
    template_name = "cases/case_report.html"

    def dispatch(self, request, *args, **kwargs):
        if request.user.role not in [User.Role.SECRETARIA, User.Role.ASESOR]:
            messages.error(request, "No tienes permiso para acceder a esta funcionalidad.")
            return redirect("dashboard_redirect")
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        cases = Case.objects.select_related(
            "beneficiary",
            "assigned_student",
            "advisor",
            "secretary"
        ).all().order_by("-created_at")

        context["user_data"] = get_topbar_user_data(self.request.user)
        context["cases"] = cases
        context["cases_count"] = cases.count()
        return context

    def post(self, request, *args, **kwargs):
        """Descarga los casos en Excel"""
        if request.POST.get("download_excel") == "true":
            return self.download_excel()
        return self.render_to_response(self.get_context_data())

    def download_excel(self):
        """Genera y descarga archivo Excel con todos los casos"""
        cases = Case.objects.select_related(
            "beneficiary",
            "assigned_student",
            "advisor",
            "secretary"
        ).all().order_by("-created_at")

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Casos"

        # Estilos
        header_fill = PatternFill(start_color="5B5CE2", end_color="5B5CE2", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")

        # Headers
        headers = [
            "ID",
            "Beneficiario",
            "Estudiante Asignado",
            "Asesor",
            "Categoría",
            "Tipo",
            "Fecha Creación",
            "Estado",
            "Etapa Actual"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        # Datos
        for row, case in enumerate(cases, 2):
            ws.cell(row=row, column=1, value=case.case_number)
            ws.cell(row=row, column=2, value=case.beneficiary.full_name if case.beneficiary else "N/A")
            ws.cell(row=row, column=3, value=case.assigned_student.full_name if case.assigned_student else "Sin asignar")
            ws.cell(row=row, column=4, value=case.advisor.full_name if case.advisor else "N/A")
            ws.cell(row=row, column=5, value=case.get_category_display() if case.category else "N/A")
            ws.cell(row=row, column=6, value=case.get_specific_type_display() or "N/A")
            ws.cell(row=row, column=7, value=case.created_at.strftime("%d/%m/%Y") if case.created_at else "N/A")
            ws.cell(row=row, column=8, value=case.get_status_display())
            ws.cell(row=row, column=9, value=case.get_current_stage_display() if case.current_stage else "N/A")

        # Ajustar ancho de columnas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(1, col).column_letter].width = 18

        # Respuesta
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="reporte_casos.xlsx"'
        wb.save(response)
        return response


class CaseDetailView(LoginRequiredMixin, TemplateView):
    template_name = "cases/case_detail.html"

    def dispatch(self, request, *args, **kwargs):
        self.case = get_object_or_404(
            Case.objects.select_related(
                "beneficiary",
                "assigned_student",
                "advisor",
                "secretary",
            ),
            id=self.kwargs["case_id"],
        )
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        history_items = (
            BitacoraEntry.objects.filter(case=self.case)
            .select_related("author")
            .order_by("-created_at")
        )
        #documents_count = Archive.objects.filter(case=self.case).count()
        stages = CaseStageManager.get_stage_display_info(self.case)

        context["user_data"] = get_topbar_user_data(self.request.user)
        context["case"] = self.case
        context["can_reassign"] = self.request.user.role == User.Role.SECRETARIA
        context["can_change_stage"] = self.request.user.role in [
            User.Role.SECRETARIA,
            User.Role.ASESOR,
        ]
        context["bitacora_count"] = history_items.count()
        context["documents_count"] = documents_count
        context["history_items"] = history_items[:5]
        context["stages"] = stages
        return context


class ReassignCaseView(LoginRequiredMixin, TemplateView):
    template_name = "cases/reassign_case.html"

    def dispatch(self, request, *args, **kwargs):
        if request.user.role != User.Role.SECRETARIA:
            messages.error(request, "No autorizado.")
            return redirect("dashboard_redirect")

        self.case = get_object_or_404(
            Case.objects.select_related(
                "beneficiary",
                "assigned_student",
                "advisor",
                "secretary",
            ),
            id=self.kwargs["case_id"],
        )
        return super().dispatch(request, *args, **kwargs)

    def get_students(self):
        return User.objects.filter(
            role=User.Role.ESTUDIANTE,
            is_active=True,
        ).order_by("first_name", "last_name")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["user_data"] = get_topbar_user_data(self.request.user)
        context["case"] = self.case
        context["students"] = self.get_students()
        return context

    def post(self, request, *args, **kwargs):
        new_student_id = request.POST.get("new_student")
        reason = request.POST.get("reason", "").strip()

        if not new_student_id:
            messages.error(request, "Debes seleccionar un nuevo estudiante.")
            return redirect("reassign_case", case_id=self.case.id)

        try:
            new_student = User.objects.get(
                id=new_student_id,
                role=User.Role.ESTUDIANTE,
                is_active=True,
            )
        except User.DoesNotExist:
            messages.error(request, "El estudiante seleccionado no es válido.")
            return redirect("reassign_case", case_id=self.case.id)

        if self.case.assigned_student == new_student:
            messages.error(
                request,
                "El caso ya está asignado a este estudiante.",
            )
            return redirect("reassign_case", case_id=self.case.id)

        CaseAssignment.objects.create(
            case=self.case,
            assigned_student=new_student,
            assigned_advisor=request.user,
            assignment_type=CaseAssignment.AssignmentType.MANUAL,
            case_category=self.case.case_type,
            notes=reason or None,
        )

        self.case.assigned_student = new_student
        self.case.status = Case.CaseStatus.ASIGNADO
        self.case.current_stage = Case.CaseStage.ASSIGNMENT
        self.case.save(
            update_fields=[
                "assigned_student",
                "status",
                "current_stage",
                "updated_at",
            ]
        )

        content = f"Caso reasignado a {getattr(new_student, 'full_name', str(new_student))}."
        if reason:
            content += f" Motivo: {reason}"

        BitacoraEntry.objects.create(
            case=self.case,
            author=request.user,
            entry_type=BitacoraEntry.EntryType.ASIGNACION,
            content=content,
        )

        messages.success(request, "Caso reasignado exitosamente.")
        return redirect("case_detail", case_id=self.case.id)


class CaseBitacoraView(LoginRequiredMixin, TemplateView):
    template_name = "cases/case_bitacora.html"

    def dispatch(self, request, *args, **kwargs):
        self.case = get_object_or_404(
            Case.objects.select_related(
                "beneficiary",
                "assigned_student",
                "advisor",
                "secretary",
            ),
            id=self.kwargs["case_id"],
        )
        return super().dispatch(request, *args, **kwargs)

    def get_entries(self):
        return (
            BitacoraEntry.objects.filter(case=self.case)
            .select_related("author")
            .order_by("-created_at")
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["user_data"] = get_topbar_user_data(self.request.user)
        context["case"] = self.case
        context["entries"] = self.get_entries()
        context["form"] = kwargs.get("form", BitacoraEntryForm())
        return context

    def post(self, request, *args, **kwargs):
        form = BitacoraEntryForm(request.POST, request.FILES)

        if form.is_valid():
            entry = form.save(commit=False)
            entry.case = self.case
            entry.author = request.user
            entry.save()

            files = form.cleaned_data.get("files", [])
            for file in files:
                #Archive.objects.create(
                    id=str(uuid.uuid4())[:15],
                    name=file.name[:100],
                    description="Archivo adjunto desde bitácora",
                    typeArchive=(getattr(file, "content_type", "unknown") or "unknown")[:15],
                    path=1,
                    case=self.case,
                #)

            messages.success(request, "Entrada agregada correctamente a la bitácora.")
            return redirect("case_bitacora", case_id=self.case.id)

        messages.error(request, "No se pudo guardar la entrada. Revisa el formulario.")
        return self.render_to_response(self.get_context_data(form=form))


class CaseDistributionListView(LoginRequiredMixin, TemplateView):
    """
    Vista para listar todos los casos SIN_ASIGNAR que pueden distribuirse (HU3).
    Acceso solo para Secretaría y Asesores.
    """
    template_name = "cases/case_distribution_list.html"

    def dispatch(self, request, *args, **kwargs):
        if request.user.role not in [User.Role.SECRETARIA, User.Role.ASESOR]:
            messages.error(
                request,
                "No tienes permiso para acceder a esta funcionalidad.",
            )
            return redirect("dashboard_redirect")
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        cases = (
            Case.objects.filter(status=Case.CaseStatus.SIN_ASIGNAR)
            .select_related("beneficiary", "assigned_student", "advisor", "secretary")
            .order_by("-created_at")
        )

        cases_with_status = []
        complete_count = 0
        for case in cases:
            completion_check = CaseCompletionValidator.is_complete(case)
            documents_info = CaseDocumentService.get_required_documents_for_display(case)
            
            # Construir información de documentos para el template
            doc_status = {
                'DOCUMENTO': {
                    'exists': documents_info.get('DOCUMENTO', {}).get('exists', False),
                    'valid': documents_info.get('DOCUMENTO', {}).get('valid', False),
                },
                'RECIBO_SERVICIOS': {
                    'exists': documents_info.get('RECIBO_SERVICIOS', {}).get('exists', False),
                    'valid': documents_info.get('RECIBO_SERVICIOS', {}).get('valid', False),
                },
                'FOTO': {
                    'exists': documents_info.get('FOTO', {}).get('exists', False),
                    'valid': documents_info.get('FOTO', {}).get('valid', False),
                },
            }
            
            cases_with_status.append(
                {
                    "case": case,
                    "is_complete": completion_check["is_complete"],
                    "completion_details": completion_check,
                    "documents": doc_status,
                }
            )
            if completion_check["is_complete"]:
                complete_count += 1

        context.update(
            {
                "user_data": get_topbar_user_data(self.request.user),
                "cases": cases_with_status,
                "casos_completos_count": complete_count,
                "students": User.objects.filter(
                    role=User.Role.ESTUDIANTE,
                    is_active=True,
                ).order_by("first_name", "last_name"),
            }
        )
        return context

    def post(self, request, *args, **kwargs):
        case_id = request.POST.get("case_id")
        action = request.POST.get("action")

        try:
            case = Case.objects.get(
                id=case_id,
                status=Case.CaseStatus.SIN_ASIGNAR,
            )
        except Case.DoesNotExist:
            return JsonResponse(
                {
                    "success": False,
                    "message": "El caso no existe o no está disponible.",
                },
                status=400,
            )

        completion_check = CaseCompletionValidator.is_complete(case)
        is_complete = completion_check["is_complete"]

        if action == "assign" and is_complete:
            return self._handle_assignment_json(request, case)
        elif action == "review" and not is_complete:
            return self._handle_send_to_review_json(request, case)
        else:
            return JsonResponse(
                {
                    "success": False,
                    "message": "Acción no válida para este caso.",
                },
                status=400,
            )

    def _handle_assignment_json(self, request, case):
        category = request.POST.get("category")
        case_type = request.POST.get("case_type")
        student_id = request.POST.get("assigned_student")

        if not all([category, case_type, student_id]):
            return JsonResponse(
                {
                    "success": False,
                    "message": "Debes completar todos los campos requeridos.",
                },
                status=400,
            )

        try:
            student = User.objects.get(
                id=student_id,
                role=User.Role.ESTUDIANTE,
                is_active=True,
            )
        except User.DoesNotExist:
            return JsonResponse(
                {
                    "success": False,
                    "message": "El estudiante seleccionado no es válido.",
                },
                status=400,
            )

        success, message, assignment = CaseDistributionService.assign_case_manually(
            case=case,
            student=student,
            category=category,
            case_type=case_type,
            assigned_by=request.user,
            notes="",
        )

        if success:
            return JsonResponse(
                {
                    "success": True,
                    "message": "¡Caso asignado exitosamente!",
                    "case_id": case.id,
                },
                status=200,
            )

        return JsonResponse(
            {
                "success": False,
                "message": message,
            },
            status=400,
        )

    def _handle_send_to_review_json(self, request, case):
        success, message = CaseDistributionService.send_to_review(
            case=case,
            sent_by=request.user,
            reason="Enviado a revisión desde el listado de distribución",
        )

        if success:
            return JsonResponse(
                {
                    "success": True,
                    "message": "Caso enviado a revisión exitosamente.",
                    "case_id": case.id,
                },
                status=200,
            )

        return JsonResponse(
            {
                "success": False,
                "message": message,
            },
            status=400,
        )


class CaseDistributionView(LoginRequiredMixin, TemplateView):
    """
    Vista para el reparto manual de casos (HU3).
    """
    template_name = "cases/case_distribution.html"

    def dispatch(self, request, *args, **kwargs):
        if request.user.role not in [User.Role.SECRETARIA, User.Role.ASESOR]:
            messages.error(
                request,
                "No tienes permiso para acceder a esta funcionalidad.",
            )
            return redirect("dashboard_redirect")

        self.case = get_object_or_404(
            Case.objects.select_related(
                "beneficiary",
                "assigned_student",
                "advisor",
                "secretary",
            ),
            id=self.kwargs["case_id"],
        )
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        if self.case.status != Case.CaseStatus.SIN_ASIGNAR:
            context["error"] = (
                f"El caso no está disponible para reparto. "
                f"Estado: {self.case.get_status_display()}"
            )
            context["case"] = self.case
            context["user_data"] = get_topbar_user_data(self.request.user)
            return context

        completion_check = CaseCompletionValidator.is_complete(self.case)
        is_complete = completion_check["is_complete"]
        documents = CaseDocumentService.get_required_documents_for_display(self.case)
        form = kwargs.get("form") or CaseDistributionForm()

        context.update(
            {
                "user_data": get_topbar_user_data(self.request.user),
                "case": self.case,
                "is_complete": is_complete,
                "completion_details": completion_check,
                "documents": documents,
                "form": form,
                "button_text": "Asignar Caso" if is_complete else "Enviar a Revisión",
                "button_enabled": is_complete,
                "button_class": "btn-primary" if is_complete else "btn-warning",
            }
        )
        return context

    def post(self, request, *args, **kwargs):
        if self.case.status != Case.CaseStatus.SIN_ASIGNAR:
            messages.error(request, "El caso no está disponible para reparto.")
            return redirect("dashboard_redirect")

        completion_check = CaseCompletionValidator.is_complete(self.case)
        if completion_check["is_complete"]:
            return self._handle_assignment(request)
        return self._handle_send_to_review(request)

    def _handle_assignment(self, request):
        form = CaseDistributionForm(request.POST)

        if form.is_valid():
            success, message, assignment = CaseDistributionService.assign_case_manually(
                case=self.case,
                student=form.cleaned_data["assigned_student"],
                category=form.cleaned_data["category"],
                case_type=form.cleaned_data["case_type"],
                assigned_by=request.user,
                notes=form.cleaned_data.get("notes", "").strip(),
            )

            if success:
                messages.success(request, "¡Caso asignado exitosamente!")
                return redirect("case_detail", case_id=self.case.id)

            messages.error(request, message)
            return self.render_to_response(self.get_context_data(form=form))

        messages.error(request, "Por favor completa todos los campos requeridos.")
        return self.render_to_response(self.get_context_data(form=form))

    def _handle_send_to_review(self, request):
        reason = request.POST.get("reason", "").strip()

        success, message = CaseDistributionService.send_to_review(
            case=self.case,
            sent_by=request.user,
            reason=reason,
        )

        if success:
            messages.success(request, "Caso enviado a revisión exitosamente.")
            return redirect("case_detail", case_id=self.case.id)

        messages.error(request, message)
        return self.render_to_response(self.get_context_data())


class AssignedCasesCategoriesView(LoginRequiredMixin, TemplateView):
    template_name = "cases/assigned_cases_categories.html"

    def dispatch(self, request, *args, **kwargs):
        if request.user.role != User.Role.SECRETARIA:
            messages.error(
                request,
                "No tienes permiso para acceder a esta funcionalidad.",
            )
            return redirect("dashboard_redirect")
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        assigned_cases = Case.objects.filter(
            status__in=[
                Case.CaseStatus.ASIGNADO,
                Case.CaseStatus.EN_PROCESO,
                Case.CaseStatus.ESPERANDO_USUARIO,
                Case.CaseStatus.EN_REVISION,
            ]
        ).select_related("beneficiary", "assigned_student")

        categories = {}
        for case in assigned_cases:
            category = case.category if case.category else "Sin categoría"
            category_label = (
                case.get_category_display()
                if hasattr(case, "get_category_display")
                else category
            )

            if category not in categories:
                categories[category] = {
                    "label": category_label,
                    "count": 0,
                    "types": {},
                }

            categories[category]["count"] += 1

            case_type_short = case.case_type_specific if case.case_type_specific else "Sin tipo"
            case_type_label = get_type_label(case_type_short)

            if case_type_label not in categories[category]["types"]:
                categories[category]["types"][case_type_label] = 0

            categories[category]["types"][case_type_label] += 1

        categories_list = [
            {
                "key": key,
                "label": value["label"],
                "count": value["count"],
                "types": sorted(value["types"].items(), key=lambda x: x[1], reverse=True),
            }
            for key, value in categories.items()
        ]
        categories_list.sort(key=lambda x: x["count"], reverse=True)

        context["user_data"] = get_topbar_user_data(self.request.user)
        context["categories"] = categories_list
        context["total_assigned"] = assigned_cases.count()
        return context


class AssignedCasesAPIView(LoginRequiredMixin, View):
    def dispatch(self, request, *args, **kwargs):
        if request.user.role != User.Role.SECRETARIA:
            return JsonResponse({"error": "No autorizado"}, status=403)
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        category = request.GET.get("category", "")
        case_type = request.GET.get("type", "")

        cases_query = Case.objects.filter(
            status__in=[
                Case.CaseStatus.ASIGNADO,
                Case.CaseStatus.EN_PROCESO,
                Case.CaseStatus.ESPERANDO_USUARIO,
                Case.CaseStatus.EN_REVISION,
            ]
        ).select_related("beneficiary", "assigned_student")

        if category and category != "Sin categoría":
            cases_query = cases_query.filter(category=category)

        if case_type and case_type != "Sin tipo":
            cases_query = cases_query.filter(case_type_specific=case_type)

        cases_data = []
        for case in cases_query.order_by("-created_at"):
            cases_data.append(
                {
                    "id": case.id,
                    "case_number": case.case_number,
                    "beneficiary_name": case.beneficiary.full_name,
                    "student_name": (
                        case.assigned_student.full_name if case.assigned_student else "Sin asignar"
                    ),
                    "status": case.get_status_display(),
                    "category": (
                        case.get_category_display()
                        if hasattr(case, "get_category_display")
                        else case.category
                    ),
                    "type": (
                        get_type_label(case.case_type_specific)
                        if case.case_type_specific
                        else "Sin tipo"
                    ),
                    "description": (
                        case.description[:100] + "..."
                        if len(case.description) > 100
                        else case.description
                    ),
                }
            )

        return JsonResponse({"cases": cases_data})


class AssignedCasesFilteredView(LoginRequiredMixin, TemplateView):
    template_name = "cases/assigned_cases_filtered.html"

    def dispatch(self, request, *args, **kwargs):
        if request.user.role != User.Role.SECRETARIA:
            messages.error(
                request,
                "No tienes permiso para acceder a esta funcionalidad.",
            )
            return redirect("dashboard_redirect")
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        category = self.request.GET.get("category", "")
        case_type_code = self.request.GET.get("type", "")

        cases_query = Case.objects.filter(
            status__in=[
                Case.CaseStatus.ASIGNADO,
                Case.CaseStatus.EN_PROCESO,
                Case.CaseStatus.ESPERANDO_USUARIO,
                Case.CaseStatus.EN_REVISION,
            ]
        ).select_related("beneficiary", "assigned_student")

        if category and category != "Sin categoría":
            cases_query = cases_query.filter(category=category)

        if case_type_code and case_type_code != "Sin tipo":
            cases_query = cases_query.filter(case_type_specific=case_type_code)

        context["user_data"] = get_topbar_user_data(self.request.user)
        context["cases"] = cases_query.order_by("-created_at")
        context["category"] = dict(Case.CaseCategory.choices).get(category, category)
        context["case_type"] = get_type_label(case_type_code) if case_type_code else ""
        context["cases_count"] = cases_query.count()
        return context


def panel_secretaria(request):
    casos_pendientes = Case.objects.filter(status=Case.CaseStatus.SIN_ASIGNAR)
    return render(request, "cases/panel_secretaria.html", {"casos": casos_pendientes})


def ejecutar_reparto(request, caso_id):
    if request.method == "POST":
        exito, mensaje = asignar_caso_automatico(caso_id, request.user)
        if exito:
            messages.success(request, mensaje)
        else:
            messages.error(request, mensaje)
    return redirect("panel_secretaria")


def calendario_seguimientos(request):
    seguimientos = BitacoraEntry.objects.all()
    return render(request, "cases/calendario.html", {"seguimientos": seguimientos})


class PendingCasesView(LoginRequiredMixin, ListView):
    model = Case
    template_name = "cases/pending_cases.html"
    context_object_name = "cases"
    paginate_by = 10

    def dispatch(self, request, *args, **kwargs):
        if request.user.role not in [User.Role.SECRETARIA, User.Role.ESTUDIANTE]:
            raise PermissionDenied
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        return (
            Case.objects.filter(status=Case.CaseStatus.SIN_ASIGNAR)
            .select_related("beneficiary")
            .order_by("-created_at")
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["total_pending"] = self.get_queryset().count()
        context["user_data"] = get_topbar_user_data(self.request.user)
        return context


class AssignedCaseDetailView(LoginRequiredMixin, DetailView):
    model = Case
    template_name = "cases/assigned_case_detail.html"
    context_object_name = "case"
    pk_url_kwarg = "case_id"

    def get_object(self, queryset=None):
        case = get_object_or_404(
            Case.objects.select_related(
                "beneficiary",
                "assigned_student",
                "advisor",
                "secretary",
            ),
            id=self.kwargs["case_id"],
        )
        user = self.request.user

        if user.role == User.Role.SECRETARIA:
            return case

        if user.role == User.Role.ESTUDIANTE:
            if case.assigned_student != user:
                raise PermissionDenied
            return case

        if user.role == User.Role.ASESOR:
            if case.advisor != user:
                raise PermissionDenied
            return case

        raise PermissionDenied

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["assignment"] = (
            CaseAssignment.objects.filter(case=self.object)
            .order_by("-assigned_at")
            .first()
        )
        context["binnacles"] = (
            self.object.bitacora_entries.select_related("author").order_by("-created_at")
        )
        context["archives"] = self.object.archives.all()
        context["user_data"] = get_topbar_user_data(self.request.user)
        return context

def reparto_automatico_view(request):
    """
    Vista para ejecutar el reparto automático equitativo de casos completos (HU3).
    """
    if request.method == "POST":
        if request.user.role not in [User.Role.SECRETARIA, User.Role.ASESOR]:
            messages.error(request, "No tienes permiso para ejecutar el reparto automático.")
            return redirect("case_distribution_list")

        # Ejecutar reparto automático masivo
        success, message = CaseAutomaticDistributionService.ejecutar_reparto_automatico_masivo(
            executed_by=request.user
        )

        if success:
            messages.success(request, message)
            return redirect("case_distribution_list")
        else:
            messages.warning(request, message)
            return redirect("case_distribution_list")
    
    return redirect("case_distribution_list")


class CaseSearchAPIView(LoginRequiredMixin, View):
    """
    API endpoint para buscar casos por número.
    Solo usuarios autenticados pueden acceder.
    """
    
    def get(self, request):
        search_term = request.GET.get('q', '').strip()
        
        if not search_term or len(search_term) < 1:
            return JsonResponse({
                "results": [],
                "message": "Ingresa un número de caso para buscar"
            })
        
        # Buscar casos por número (case-insensitive)
        cases = Case.objects.filter(
            case_number__icontains=search_term
        ).select_related(
            'beneficiary',
            'assigned_student'
        ).order_by('-created_at')[:10]  # Máximo 10 resultados
        
        results = []
        for case in cases:
            results.append({
                'id': case.id,
                'case_number': case.case_number,
                'beneficiary_name': case.beneficiary.full_name if case.beneficiary else 'Sin beneficiario',
                'student_name': case.assigned_student.full_name if case.assigned_student else 'Sin asignar',
                'status': case.status,
                'status_display': case.get_status_display(),
            })
        
        return JsonResponse({
            "results": results,
            "count": len(results)
        })


class CaseSearchByIdAPIView(LoginRequiredMixin, View):
    """
    API endpoint para buscar un caso por ID exacto (5 dígitos numéricos).
    Busca TODOS los casos registrados (asignados y sin asignar).
    """
    
    def get(self, request):
        case_id = request.GET.get('id', '').strip()
        
        # Validar que sea exactamente 5 dígitos
        if not case_id or len(case_id) != 5 or not case_id.isdigit():
            return JsonResponse({
                "case": None,
                "error": "Ingresa un ID válido de 5 dígitos numéricos"
            }, status=400)
        
        # Buscar por ID exacto
        try:
            case = Case.objects.select_related(
                'beneficiary',
                'assigned_student'
            ).get(case_number=case_id)
            
            return JsonResponse({
                "case": {
                    'id': case.id,
                    'sequence_number': case.sequence_number,
                    'case_number': case.case_number,
                    'beneficiary_name': case.beneficiary.full_name if case.beneficiary else 'Sin beneficiario',
                    'student_name': case.assigned_student.full_name if case.assigned_student else 'Sin asignar',
                    'status': case.status,
                    'status_display': case.get_status_display(),
                    'created_at': case.created_at.strftime('%d/%m/%Y'),
                },
                "found": True
            })
        except Case.DoesNotExist:
            return JsonResponse({
                "case": None,
                "found": False,
                "error": f"No se encuentra un caso con ese ID"
            }, status=404)


class CaseSearchUnassignedByIdAPIView(LoginRequiredMixin, View):
    """
    API endpoint para buscar un caso SIN_ASIGNAR por ID exacto (5 dígitos numéricos).
    Usado para la barra de búsqueda en case_distribution_list (reparto de casos).
    """
    
    def get(self, request):
        case_id = request.GET.get('id', '').strip()
        
        # Validar que sea exactamente 5 dígitos
        if not case_id or len(case_id) != 5 or not case_id.isdigit():
            return JsonResponse({
                "case": None,
                "error": "Ingresa un ID válido de 5 dígitos numéricos"
            }, status=400)
        
        # Buscar por ID exacto entre casos SIN_ASIGNAR
        try:
            case = Case.objects.filter(
                status=Case.CaseStatus.SIN_ASIGNAR
            ).select_related(
                'beneficiary',
                'assigned_student'
            ).get(case_number=case_id)
            
            # Obtener información de completitud
            completion_check = CaseCompletionValidator.is_complete(case)
            
            return JsonResponse({
                "case": {
                    'id': case.id,
                    'sequence_number': case.sequence_number,
                    'case_number': case.case_number,
                    'beneficiary_name': case.beneficiary.full_name if case.beneficiary else 'Sin beneficiario',
                    'student_name': case.assigned_student.full_name if case.assigned_student else 'Sin asignar',
                    'status': case.status,
                    'status_display': case.get_status_display(),
                    'created_at': case.created_at.strftime('%d/%m/%Y'),
                    'is_complete': completion_check['is_complete'],
                },
                "found": True
            })
        except Case.DoesNotExist:
            return JsonResponse({
                "case": None,
                "found": False,
                "error": f"No se encuentra un caso con ese ID"
            }, status=404)


class CaseClosureAPIView(LoginRequiredMixin, View):
    """
    API endpoint para cerrar un caso y registrar el tipo de cierre.
    Usado cuando se cambia el estado a CERRADO en case_detail.
    Solo SECRETARIA y ASESOR pueden cerrar casos.
    """
    
    def post(self, request):
        # Validar permisos
        if request.user.role not in [User.Role.SECRETARIA, User.Role.ASESOR]:
            return JsonResponse({
                "success": False,
                "error": "No tienes permiso para cerrar casos"
            }, status=403)
        
        case_id = request.POST.get('case_id')
        closure_type = request.POST.get('closure_type')
        
        if not case_id or not closure_type:
            return JsonResponse({
                "success": False,
                "error": "Faltan parámetros requeridos"
            }, status=400)
        
        if closure_type not in [choice[0] for choice in Case.ClosureType.choices]:
            return JsonResponse({
                "success": False,
                "error": "Tipo de cierre inválido"
            }, status=400)
        
        try:
            case = Case.objects.get(id=case_id)
            
            # Cambiar estado a CERRADO y guardar tipo de cierre
            case.status = Case.CaseStatus.CERRADO
            case.closure_type = closure_type
            case.current_stage = CaseStageManager.calculate_current_stage(case)
            case.save(update_fields=['status', 'closure_type', 'current_stage', 'updated_at'])
            
            # Crear entrada en bitácora
            closure_label = dict(Case.ClosureType.choices).get(closure_type, "Desconocido")
            BitacoraEntry.objects.create(
                case=case,
                author=request.user,
                entry_type=BitacoraEntry.EntryType.ACTUALIZACION,
                content=f"Caso cerrado: {closure_label}"
            )
            
            return JsonResponse({
                "success": True,
                "message": "Caso cerrado exitosamente"
            })
        except Case.DoesNotExist:
            return JsonResponse({
                "success": False,
                "error": "Caso no encontrado"
            }, status=404)
        except Exception as e:
            return JsonResponse({
                "success": False,
                "error": str(e)
            }, status=500)


class CaseStatusChangeAPIView(LoginRequiredMixin, View):
    """
    API endpoint para cambiar el estado de un caso (sin cierre).
    Solo SECRETARIA y ASESOR pueden cambiar el estado.
    Sincroniza automáticamente las etapas del caso con el nuevo estado.
    """

    # Mapa: estado → etapa correspondiente
    STATUS_TO_STAGE = {
        Case.CaseStatus.SIN_ASIGNAR:       Case.CaseStage.UNASSIGNED,           # 0
        Case.CaseStatus.ASIGNADO:          Case.CaseStage.ASSIGNMENT,            # 1
        Case.CaseStatus.EN_PROCESO:        Case.CaseStage.INFORMATION_GATHERING, # 2
        Case.CaseStatus.ESPERANDO_USUARIO: Case.CaseStage.ANALYSIS_DRAFTING,     # 3
        Case.CaseStatus.EN_REVISION:       Case.CaseStage.SUPERVISOR_REVIEW,     # 4
        Case.CaseStatus.DOCUMENTACION:     Case.CaseStage.UNASSIGNED,            # 0 (doc pendiente)
    }

    def post(self, request):
        # Validar permisos
        if request.user.role not in [User.Role.SECRETARIA, User.Role.ASESOR]:
            return JsonResponse({
                "success": False,
                "error": "No tienes permiso para cambiar el estado del caso"
            }, status=403)

        case_id = request.POST.get('case_id')
        new_status = request.POST.get('status')

        if not case_id or not new_status:
            return JsonResponse({
                "success": False,
                "error": "Faltan parámetros requeridos"
            }, status=400)

        valid_statuses = [choice[0] for choice in Case.CaseStatus.choices]
        if new_status not in valid_statuses:
            return JsonResponse({
                "success": False,
                "error": "Estado inválido"
            }, status=400)

        try:
            case = Case.objects.get(id=case_id)

            # Determinar la nueva etapa según el estado
            new_stage = self.STATUS_TO_STAGE.get(new_status, Case.CaseStage.UNASSIGNED)

            # Aplicar cambios
            case.status = new_status
            case.current_stage = new_stage
            case.save(update_fields=['status', 'current_stage', 'updated_at'])

            # Crear entrada en bitácora
            status_label = dict(Case.CaseStatus.choices).get(new_status, "Desconocido")
            stage_label = dict(Case.CaseStage.choices).get(new_stage, "")
            BitacoraEntry.objects.create(
                case=case,
                author=request.user,
                entry_type=BitacoraEntry.EntryType.ACTUALIZACION,
                content=f"Estado del caso cambiado a: {status_label}. Etapa actual: {stage_label}"
            )

            mensaje = f'El caso "{case.case_number}" cambió a {status_label.lower()}'

            return JsonResponse({
                "success": True,
                "message": mensaje,
                "new_stage": new_stage,
                "new_stage_label": stage_label,
            })

        except Case.DoesNotExist:
            return JsonResponse({
                "success": False,
                "error": "Caso no encontrado"
            }, status=404)
        except Exception as e:
            return JsonResponse({
                "success": False,
                "error": str(e)
            }, status=500)